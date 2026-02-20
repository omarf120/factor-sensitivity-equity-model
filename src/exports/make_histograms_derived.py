#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# -----------------------------
# Config
# -----------------------------
YEARS = [2004, 2008, 2014, 2020, 2024]

# Repo-relative paths
REPO_ROOT = Path(__file__).resolve().parents[2]
PANEL_PATH = REPO_ROOT / "data" / "processed" / "final_factor_panel_raw_clean.csv"
TEMPLATE_PATH = REPO_ROOT / "configs" / "templates" / "histogram_template.xlsx"
OUT_DIR = REPO_ROOT / "outputs" / "tables" / "histograms_derived"  # keep same folder (or change if you want)

# Template placement
TITLE_CELL = "B1"
START_CELL = "E11"

# -----------------------------
# Variables
# -----------------------------
# Original 11 requested series (deliverable names)
FUNDAMENTAL_VARS = [
    "book_value_per_share",          # from bv_ps
    "earnings_per_share",            # from epspxq
    "cf_per_share",                  # from cf_ps
    "sales_revenue_q_m",             # from saleq
    "total_assets_q_m",              # from atq
    "total_liabilities_q_m",         # from ltq
    "shareholders_equity_q_m",       # from seqq
    "shares_outstanding_q_t",        # from cshoq
    "sales_per_share",               # from sales_ps
]

WEEKLY_VARS = [
    "weekly_log_return",             # from weekly_log_ret_total
    "price",                         # from adj_prc
]

# New 7 ratio vars
RATIO_VARS = ["bvp", "ep", "cfp", "sp", "roa", "roe", "cf_sales"]

# Treat ratios like fundamentals for histogramming (avoid weekly repetition)
RATIO_AS_FUNDAMENTAL = True


def excel_safe_name(s: str) -> str:
    """Filename-safe (Windows/macOS)."""
    bad = '<>:"/\\|?*'
    for ch in bad:
        s = s.replace(ch, "_")
    return s.strip().replace(" ", "_").lower()


def require_columns(df: pd.DataFrame, cols: list[str], label: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise KeyError(
            f"Missing required columns in {label}:\n"
            + "\n".join(f"  - {c}" for c in missing)
        )


def write_series_to_template(series: pd.Series, var_name: str, year: int, out_path: Path) -> None:
    """Load fresh template, write title in B1, write data starting at E11."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    ws[TITLE_CELL].value = f"{year} - {var_name}"

    start_row = ws[START_CELL].row
    start_col = ws[START_CELL].column

    values = series.tolist()
    for i, v in enumerate(values):
        if pd.isna(v):
            v = None
        ws.cell(row=start_row + i, column=start_col, value=v)

    wb.save(out_path)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not PANEL_PATH.exists():
        raise FileNotFoundError(f"Clean panel not found: {PANEL_PATH}")

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Histogram template not found: {TEMPLATE_PATH}")

    df = pd.read_csv(PANEL_PATH, low_memory=False)
    df.columns = [c.strip() for c in df.columns]

    # Need these for year filtering + dedupe
    base_required = ["PERMNO", "date", "datadate"]
    require_columns(df, base_required, PANEL_PATH.name)

    # --- Build the original 11 series from your cleaned panel columns ---
    # Require the raw building-block cols used for these series:
    raw_required = ["bv_ps", "epspxq", "cf_ps", "saleq", "atq", "ltq", "seqq", "cshoq", "sales_ps", "weekly_log_ret_total", "adj_prc"]
    require_columns(df, raw_required, PANEL_PATH.name)

    # Derive the same 11 deliverable names as before
    df["book_value_per_share"] = df["bv_ps"]
    df["earnings_per_share"] = df["epspxq"]
    df["cf_per_share"] = df["cf_ps"]
    df["sales_revenue_q_m"] = df["saleq"]
    df["total_assets_q_m"] = df["atq"]
    df["total_liabilities_q_m"] = df["ltq"]
    df["shareholders_equity_q_m"] = df["seqq"]
    df["shares_outstanding_q_t"] = df["cshoq"]
    df["sales_per_share"] = df["sales_ps"]
    df["weekly_log_return"] = df["weekly_log_ret_total"]
    df["price"] = df["adj_prc"]

    # Require ratio cols too (for the extra 35)
    require_columns(df, RATIO_VARS, PANEL_PATH.name)

    # Parse dates
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["datadate"] = pd.to_datetime(df["datadate"], errors="coerce")
    df = df.dropna(subset=["date"])

    all_fundamental_like = set(FUNDAMENTAL_VARS)
    if RATIO_AS_FUNDAMENTAL:
        all_fundamental_like |= set(RATIO_VARS)

    all_weekly_like = set(WEEKLY_VARS)
    if not RATIO_AS_FUNDAMENTAL:
        all_weekly_like |= set(RATIO_VARS)

    all_vars = sorted(list(all_fundamental_like | all_weekly_like))  # 16 total if ratios fundamental-like, else 18

    n_written = 0
    for year in YEARS:
        df_y = df[df["date"].dt.year == year].copy()
        if df_y.empty:
            print(f"[SKIP] No rows for year={year}")
            continue

        # Dedupe quarterly-style variables by firm-quarter
        df_funda_y = df_y.dropna(subset=["datadate"]).drop_duplicates(subset=["PERMNO", "datadate"])

        for var in all_vars:
            if var in all_fundamental_like:
                s = df_funda_y[var]
            else:
                s = df_y[var]

            # numeric only
            s = pd.to_numeric(s, errors="coerce").dropna()

            out_name = f"{excel_safe_name(var)}_{year}_hist.xlsx"
            out_path = OUT_DIR / out_name

            write_series_to_template(s, var_name=var, year=year, out_path=out_path)
            n_written += 1
            print(f"[OK] {year} | {var} | n={len(s):,} -> {out_path.relative_to(REPO_ROOT)}")

    print(f"\nDone. Wrote {n_written} histogram workbooks to: {OUT_DIR.relative_to(REPO_ROOT)}")
    # Expected: 55 + 35 = 90 when all 5 years exist


if __name__ == "__main__":
    main()