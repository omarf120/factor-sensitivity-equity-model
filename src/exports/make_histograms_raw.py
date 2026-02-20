#!/usr/bin/env python3
from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


# -----------------------------
# Config
# -----------------------------
YEARS = [2004, 2008, 2014, 2020, 2024]

# Repo-relative paths
REPO_ROOT = Path(__file__).resolve().parents[2]
PANEL_PATH = REPO_ROOT / "data" / "processed" / "final_factor_panel_raw.csv"
TEMPLATE_PATH = REPO_ROOT / "configs" / "templates" / "histogram_template.xlsx"
OUT_DIR = REPO_ROOT / "outputs" / "tables" / "histograms_raw"


def excel_safe_name(s: str) -> str:
    """Filename-safe (Windows/macOS)."""
    bad = '<>:"/\\|?*'
    for ch in bad:
        s = s.replace(ch, "_")
    return s.strip().replace(" ", "_").lower()


def write_series_to_template(series: pd.Series, var_name: str, year: int, out_path: Path) -> None:
    """
    Load fresh template, write title in B1, write data starting at E11.
    """
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ---- Title in B1 ----
    ws["B1"].value = f"{year} - {var_name}"

    # ---- Data starts at E11 ----
    start_row = ws["E11"].row
    start_col = ws["E11"].column

    values = series.tolist()

    for i, v in enumerate(values):
        if pd.isna(v):
            v = None
        ws.cell(row=start_row + i, column=start_col, value=v)

    wb.save(out_path)


def require_columns(df: pd.DataFrame, cols: list[str]) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise KeyError(
            "Missing required columns in final_factor_panel.csv:\n"
            + "\n".join(f"  - {c}" for c in missing)
        )


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not PANEL_PATH.exists():
        raise FileNotFoundError(f"Processed panel not found: {PANEL_PATH}")

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Histogram template not found: {TEMPLATE_PATH}")

    df = pd.read_csv(PANEL_PATH, low_memory=False)

    # Normalize column names
    df.columns = [c.strip() for c in df.columns]

    # Required base columns from pipeline panel
    required = [
        "PERMNO", "date", "datadate",
        "adj_prc", "weekly_log_ret_total",
        "bv_ps", "saleq", "atq", "ltq", "seqq", "cshoq", "epspxq", "cf_ps", "sales_ps", "niq"
    ]
    require_columns(df, required)

    # Parse dates
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["datadate"] = pd.to_datetime(df["datadate"], errors="coerce")
    df = df.dropna(subset=["date"])  # need year filter

    # Compute the 11 requested series (deliverable names)
    df["book_value_per_share"] = df["bv_ps"]
    df["earnings_per_share"] = df["epspxq"]
    df["cf_per_share"] = df["cf_ps"]
    df["sales_revenue_q_m"] = df["saleq"]
    df["total_assets_q_m"] = df["atq"]
    df["total_liabilities_q_m"] = df["ltq"]
    df["shareholders_equity_q_m"] = df["seqq"]
    df["shares_outstanding_q_t"] = df["cshoq"]
    df["sales_per_share"] = df["sales_ps"]
    df["weekly_log_return"] = df["weekly_log_ret_total"]
    df["price"] = df["adj_prc"]
    df["net_income"] = df["niq"]

    # Which are fundamentals (dedupe by permno+datadate before histogramming)
    fundamental_vars = {
        "book_value_per_share",
        "earnings_per_share",
        "cf_per_share",
        "sales_revenue_q_m",
        "total_assets_q_m",
        "total_liabilities_q_m",
        "shareholders_equity_q_m",
        "shares_outstanding_q_t",
        "sales_per_share",
        "net_income"
    }
    weekly_vars = {"weekly_log_return", "price"}

    all_vars = list(fundamental_vars) + list(weekly_vars)
    all_vars = sorted(all_vars)  # stable order

    # Build 55 workbooks
    n_written = 0
    for year in YEARS:
        df_y = df[df["date"].dt.year == year].copy()
        if df_y.empty:
            print(f"[SKIP] No rows for year={year}")
            continue

        # Dedupe fundamentals by firm-quarter (prevents weekly repetition bias)
        df_funda_y = df_y.dropna(subset=["datadate"]).drop_duplicates(subset=["PERMNO", "datadate"])

        for var in all_vars:
            if var in fundamental_vars:
                s = df_funda_y[var]
            else:
                s = df_y[var]

            # Keep numeric only; coerce strings to NaN
            s = pd.to_numeric(s, errors="coerce").dropna()

            out_name = f"{excel_safe_name(var)}_{year}_hist.xlsx"
            out_path = OUT_DIR / out_name

            write_series_to_template(s, var_name=var, year=year, out_path=out_path)
            n_written += 1
            print(f"[OK] {year} | {var} | n={len(s):,} -> {out_path.relative_to(REPO_ROOT)}")

    print(f"\nDone. Wrote {n_written} histogram workbooks to: {OUT_DIR.relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    main()